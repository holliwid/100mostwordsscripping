import requests
from bs4 import BeautifulSoup
import re
import openpyxl

URL = 'https://studynow.ru/dicta/allwords'
page = requests.get(URL)


soup = BeautifulSoup(page.content, 'html.parser')
lis = []
r = soup.find_all(id='wordlist')
for r in r:
    result = r.find_all('td')
list_of_raw_data = re.findall('<td>(.*?)</td>',str(result))

eng_words = []
rus_words = []

flag = 0
for word in list_of_raw_data:
    if flag % 3 == 1:
        eng_words.append(word.split(',')[0])
    if flag % 3 == 2:
        rus_words.append(word.split(',')[0].split('(')[0])
    flag += 1
    if flag == 301:
        break
flag = 1
for x in eng_words:
    print(flag,x)
    flag += 1
for x in rus_words:
    print(x)


wb = openpyxl.load_workbook(filename=r'C:\Users\markovDD\OneDrive\Рабочий стол\Офисы\Dictionary.xlsx')
ws = wb.worksheets[1]
for index in range(100):
    ws.cell(row=index+7, column=3).value = rus_words[index]
    ws.cell(row=index+7, column=12).value = rus_words[index]
    ws.cell(row=index+7, column=4).value = eng_words[index]
    ws.cell(row=index+7, column=11).value = eng_words[index]
wb.save('Dictionary.xlsx')
